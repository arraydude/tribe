#!/usr/bin/env python3
"""Export Excel data to JSON for the Tribe dashboard."""

import json
import os
import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'Seguimiento compras Tribe.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'orders.json')

COLUMNS = {
    1: 'stock',
    2: 'cliente',
    3: 'item',
    4: 'cantidad',
    6: 'valorPresupuestado',
    7: 'fechaCompra',
    8: 'valorCompra',
    9: 'valorDebitado',
    10: 'tax',
    11: 'costoEnvio',
    12: 'peso',
    13: 'status',
    14: 'saldado',
    15: 'asignado',
    16: 'ganancia',
    17: 'saldadoA',
    20: 'observaciones',
}

def export():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['SEGUIMIENTO COMPRAS TRIBE']

    orders = []
    for row_num in range(2, ws.max_row + 1):
        item_val = ws.cell(row=row_num, column=3).value
        if not item_val:
            continue

        order = {'id': row_num - 1}
        for col, key in COLUMNS.items():
            val = ws.cell(row=row_num, column=col).value
            if val is None:
                order[key] = None
            elif hasattr(val, 'isoformat'):
                order[key] = val.isoformat()[:10]
            elif isinstance(val, (int, float)):
                order[key] = round(val, 2)
            elif isinstance(val, str) and '#' in val:
                order[key] = None
            else:
                order[key] = str(val).strip()

        orders.append(order)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(orders, f, ensure_ascii=False, indent=2)

    print(f'Exported {len(orders)} orders to {OUTPUT_PATH}')

if __name__ == '__main__':
    export()
